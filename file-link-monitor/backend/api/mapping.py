"""
自定义名称映射管理API
"""
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import func, or_, and_, distinct
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime
import logging
import shutil
import subprocess
import uuid
import os
from pathlib import Path

from backend.models import CustomNameMapping, LinkRecord, get_session
from backend.utils.linker import FileLinker
from backend.utils.obfuscator import FolderObfuscator

router = APIRouter()
logger = logging.getLogger(__name__)

def get_db():
    """依赖注入：获取数据库会话"""
    from backend.main import db_engine
    session = get_session(db_engine)
    try:
        yield session
    finally:
        session.close()


class MappingCreate(BaseModel):
    """创建映射请求"""
    original_name: str
    quark_name: Optional[str] = None
    baidu_name: Optional[str] = None
    xunlei_name: Optional[str] = None
    note: Optional[str] = None
    baidu_link: Optional[str] = None
    quark_link: Optional[str] = None
    xunlei_link: Optional[str] = None


class MappingUpdate(BaseModel):
    """更新映射请求"""
    quark_name: Optional[str] = None
    baidu_name: Optional[str] = None
    xunlei_name: Optional[str] = None
    enabled: Optional[bool] = None
    is_completed: Optional[bool] = None
    note: Optional[str] = None
    baidu_link: Optional[str] = None
    quark_link: Optional[str] = None
    xunlei_link: Optional[str] = None


@router.get("/mappings")
async def get_mappings(
    page: int = 1,
    page_size: int = 20,
    enabled: Optional[bool] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    获取所有自定义名称映射（分页）
    
    Args:
        page: 页码（从1开始）
        page_size: 每页数量
        enabled: 过滤启用状态
        search: 搜索原名或自定义名
    """
    try:
        query = db.query(CustomNameMapping)
        
        if enabled is not None:
            query = query.filter(CustomNameMapping.enabled == enabled)
        
        if search:
            query = query.filter(
                (CustomNameMapping.original_name.like(f'%{search}%')) |
                (CustomNameMapping.quark_name.like(f'%{search}%')) |
                (CustomNameMapping.baidu_name.like(f'%{search}%')) |
                (CustomNameMapping.xunlei_name.like(f'%{search}%'))
            )
        
        # 总数
        total = query.count()
        
        # 分页
        mappings = query.order_by(CustomNameMapping.created_at.desc())\
            .offset((page - 1) * page_size)\
            .limit(page_size)\
            .all()
        
        return {
            "success": True,
            "data": [
                {
                    "id": m.id,
                    "original_name": m.original_name,
                    "quark_name": m.quark_name,
                    "baidu_name": m.baidu_name,
                    "xunlei_name": m.xunlei_name,
                    "enabled": m.enabled,
                    "is_completed": m.is_completed,
                    "note": m.note,
                    "baidu_link": m.baidu_link,
                    "quark_link": m.quark_link,
                    "xunlei_link": m.xunlei_link,
                    "created_at": m.created_at.isoformat() if m.created_at else None,
                    "updated_at": m.updated_at.isoformat() if m.updated_at else None
                }
                for m in mappings
            ],
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (total + page_size - 1) // page_size
        }
    except Exception as e:
        logger.error(f"获取映射失败: {e}")
        return {"success": False, "message": str(e)}


@router.post("/mappings")
async def create_mapping(mapping: MappingCreate, db: Session = Depends(get_db)):
    """
    创建自定义名称映射
    
    Args:
        mapping: 映射数据
    """
    try:
        # 检查是否已存在
        existing = db.query(CustomNameMapping).filter(
            CustomNameMapping.original_name == mapping.original_name
        ).first()
        
        if existing:
            return {
                "success": False,
                "message": f"映射已存在: {mapping.original_name}"
            }
        
        # 创建映射（自动去除前后空格）
        new_mapping = CustomNameMapping(
            original_name=mapping.original_name,
            quark_name=mapping.quark_name.strip() if mapping.quark_name else None,
            baidu_name=mapping.baidu_name.strip() if mapping.baidu_name else None,
            xunlei_name=mapping.xunlei_name.strip() if mapping.xunlei_name else None,
            note=mapping.note,
            baidu_link=mapping.baidu_link,
            quark_link=mapping.quark_link,
            xunlei_link=mapping.xunlei_link
        )
        db.add(new_mapping)
        db.commit()
        db.refresh(new_mapping)
        
        logger.info(f"✅ 创建映射: {mapping.original_name} (夸克:{mapping.quark_name}, 百度:{mapping.baidu_name})")
        
        return {
            "success": True,
            "message": "映射创建成功",
            "data": {
                "id": new_mapping.id,
                "original_name": new_mapping.original_name,
                "quark_name": new_mapping.quark_name,
                "baidu_name": new_mapping.baidu_name,
                "enabled": new_mapping.enabled,
                "note": new_mapping.note
            }
        }
    except Exception as e:
        db.rollback()
        logger.error(f"创建映射失败: {e}")
        return {"success": False, "message": str(e)}


@router.put("/mappings/{mapping_id}")
async def update_mapping(
    mapping_id: int,
    mapping: MappingUpdate,
    db: Session = Depends(get_db)
):
    """
    更新自定义名称映射
    
    Args:
        mapping_id: 映射ID
        mapping: 更新数据
    """
    try:
        existing = db.query(CustomNameMapping).filter(
            CustomNameMapping.id == mapping_id
        ).first()
        
        if not existing:
            return {"success": False, "message": "映射不存在"}
        
        # 更新字段（自动去除前后空格）
        if mapping.quark_name is not None:
            existing.quark_name = mapping.quark_name.strip() if mapping.quark_name else None
        if mapping.baidu_name is not None:
            existing.baidu_name = mapping.baidu_name.strip() if mapping.baidu_name else None
        if mapping.xunlei_name is not None:
            existing.xunlei_name = mapping.xunlei_name.strip() if mapping.xunlei_name else None
        if mapping.enabled is not None:
            existing.enabled = mapping.enabled
        if mapping.is_completed is not None:
            existing.is_completed = mapping.is_completed
        if mapping.note is not None:
            existing.note = mapping.note
        if mapping.baidu_link is not None:
            existing.baidu_link = mapping.baidu_link
        if mapping.quark_link is not None:
            existing.quark_link = mapping.quark_link
        if mapping.xunlei_link is not None:
            existing.xunlei_link = mapping.xunlei_link
        
        db.commit()
        db.refresh(existing)
        
        logger.info(f"✅ 更新映射: {existing.original_name} (夸克:{existing.quark_name}, 百度:{existing.baidu_name})")
        
        return {
            "success": True,
            "message": "映射更新成功",
            "data": {
                "id": existing.id,
                "original_name": existing.original_name,
                "quark_name": existing.quark_name,
                "baidu_name": existing.baidu_name,
                "enabled": existing.enabled,
                "note": existing.note
            }
        }
    except Exception as e:
        db.rollback()
        logger.error(f"更新映射失败: {e}")
        return {"success": False, "message": str(e)}


@router.delete("/mappings/{mapping_id}")
async def delete_mapping(mapping_id: int, db: Session = Depends(get_db)):
    """
    删除自定义名称映射
    
    Args:
        mapping_id: 映射ID
    """
    try:
        mapping = db.query(CustomNameMapping).filter(
            CustomNameMapping.id == mapping_id
        ).first()
        
        if not mapping:
            return {"success": False, "message": "映射不存在"}
        
        original_name = mapping.original_name
        db.delete(mapping)
        db.commit()
        
        logger.info(f"✅ 删除映射: {original_name}")
        
        return {
            "success": True,
            "message": "映射删除成功"
        }
    except Exception as e:
        db.rollback()
        logger.error(f"删除映射失败: {e}")
        return {"success": False, "message": str(e)}


@router.get("/export/mappings")
async def export_mappings(
    enabled: Optional[bool] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    导出名称映射为Excel文件（只包含原名和自定义名两列）
    
    Args:
        enabled: 过滤启用状态
        search: 搜索关键词
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        query = db.query(CustomNameMapping)
        
        if enabled is not None:
            query = query.filter(CustomNameMapping.enabled == enabled)
        
        if search:
            query = query.filter(
                (CustomNameMapping.original_name.like(f'%{search}%')) |
                (CustomNameMapping.quark_name.like(f'%{search}%')) |
                (CustomNameMapping.baidu_name.like(f'%{search}%')) |
                (CustomNameMapping.xunlei_name.like(f'%{search}%'))
            )
        
        mappings = query.order_by(CustomNameMapping.original_name).all()
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "名称映射"
        
        # 设置表头
        headers = ["原名", "夸克显示名", "百度显示名", "迅雷显示名", "百度链接", "夸克链接", "迅雷链接"]
        ws.append(headers)
        
        # 设置表头样式
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=14)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 填充数据
        row_num = 2
        for mapping in mappings:
            ws.append([
                mapping.original_name,
                mapping.quark_name or '-',
                mapping.baidu_name or '-',
                mapping.xunlei_name or '-',
                mapping.baidu_link or '-',
                mapping.quark_link or '-',
                mapping.xunlei_link or '-'
            ])
            
            # 设置对齐和字体
            for col_num in range(1, 8):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(size=12)
            
            row_num += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 60
        ws.column_dimensions['F'].width = 60
        ws.column_dimensions['G'].width = 60
        
        # 添加统计信息
        ws.append([])
        ws.append([f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.append([f"总计: {len(mappings)} 条映射"])
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        filename = f"name_mappings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        logger.info(f"✅ 成功导出名称映射: {len(mappings)} 条")
        
        # 返回Excel文件
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except ImportError:
        logger.error("openpyxl库未安装，无法导出Excel")
        return {
            "success": False,
            "message": "openpyxl库未安装，请安装: pip install openpyxl"
        }
    except Exception as e:
        logger.error(f"导出失败: {e}")
        return {
            "success": False,
            "message": f"导出失败: {str(e)}"
        }


def process_anti_ban(target_dir: Path) -> int:
    """
    防失效处理：修改目录下所有视频文件的Hash
    
    Args:
        target_dir: 目标目录路径
        
    Returns:
        处理的文件数量
    """
    from backend.utils.obfuscator import FolderObfuscator
    
    processed_count = 0
    
    # 遍历目录下所有文件
    for file_path in target_dir.rglob('*'):
        if not file_path.is_file():
            continue
        
        # 只处理视频文件
        if not FolderObfuscator.is_video_file(str(file_path)):
            continue
        
        # 检查是否为硬链接
        stat_info = os.stat(file_path)
        if stat_info.st_nlink <= 1:
            # 不是硬链接，跳过
            logger.info(f"跳过非硬链接文件: {file_path.name}")
            continue
        
        # 生成临时文件路径
        temp_file = file_path.parent / f".{file_path.name}.tmp"
        
        try:
            # 使用ffmpeg修改Hash
            cmd = [
                'ffmpeg',
                '-i', str(file_path),
                '-map', '0',
                '-c', 'copy',
                '-metadata', f'comment={uuid.uuid4()}',
                '-metadata', f'title={uuid.uuid4()}',
                str(temp_file),
                '-y'  # 覆盖已存在的文件
            ]
            
            logger.info(f"正在处理: {file_path.name}")
            
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=300  # 5分钟超时
            )
            
            if result.returncode != 0:
                logger.error(f"ffmpeg处理失败 {file_path.name}: {result.stderr}")
                # 清理临时文件
                if temp_file.exists():
                    temp_file.unlink()
                continue
            
            # 删除原硬链接
            file_path.unlink()
            
            # 重命名临时文件
            temp_file.rename(file_path)
            
            processed_count += 1
            logger.info(f"防失效处理完成: {file_path.name}")
            
        except subprocess.TimeoutExpired:
            logger.error(f"处理超时: {file_path.name}")
            if temp_file.exists():
                temp_file.unlink()
        except Exception as e:
            logger.error(f"处理失败 {file_path.name}: {e}")
            if temp_file.exists():
                temp_file.unlink()
    
    return processed_count


class ResyncRequest(BaseModel):
    """重转请求"""
    original_name: str
    target_type: str  # 'quark', 'baidu' 或 'xunlei'


@router.post("/mappings/resync")
async def resync_to_target(
    request: ResyncRequest,
    db: Session = Depends(get_db)
):
    """
    重转剧集到指定网盘
    1. 查询该剧集的所有文件记录
    2. 查询映射表获取显示名称
    3. 删除目标网盘的文件
    4. 用新名称重新硬链接
    """
    try:
        from pathlib import Path
        import shutil
        from backend.utils.linker import FileLinker
        from backend.main import db_engine, config
        
        # 查询映射获取显示名称
        mapping = db.query(CustomNameMapping).filter(
            CustomNameMapping.original_name == request.original_name
        ).first()
        
        if not mapping:
            return {
                "success": False,
                "message": f"未找到 '{request.original_name}' 的映射配置"
            }
        
        # 清空对应网盘的链接字段
        if request.target_type == 'quark':
            mapping.quark_link = None
            logger.info(f"✅ 已清空 '{request.original_name}' 的夸克链接")
        elif request.target_type == 'baidu':
            mapping.baidu_link = None
            logger.info(f"✅ 已清空 '{request.original_name}' 的百度链接")
        elif request.target_type == 'xunlei':
            mapping.xunlei_link = None
            logger.info(f"✅ 已清空 '{request.original_name}' 的迅雷链接")
        db.commit()
        
        # 获取显示名称
        if request.target_type == 'quark':
            display_name = mapping.quark_name
        elif request.target_type == 'baidu':
            display_name = mapping.baidu_name
        elif request.target_type == 'xunlei':
            display_name = mapping.xunlei_name
        else:
            display_name = None
        
        if not display_name:
            display_name = request.original_name
        
        # 查询所有该剧集的记录
        records = db.query(LinkRecord).filter(
            LinkRecord.original_name == request.original_name
        ).all()
        
        if not records:
            return {
                "success": False,
                "message": f"未找到 '{request.original_name}' 的同步记录"
            }
        
        # 获取配置
        monitors = config.get('monitors', [])
        if not monitors:
            return {"success": False, "message": "未找到监控配置"}
        
        monitor = monitors[0]
        source_path = Path(monitor.get('source', ''))
        targets = monitor.get('targets', [])
        
        if len(targets) < 2:
            return {"success": False, "message": "目标配置不足"}
        
        # 确定目标路径
        if request.target_type == 'quark':
            target_idx = 0
        elif request.target_type == 'baidu':
            target_idx = 1
        elif request.target_type == 'xunlei':
            target_idx = 2
        else:
            return {"success": False, "message": "不支持的网盘类型"}
        
        if len(targets) <= target_idx:
            return {"success": False, "message": f"目标配置不足，需要至少{target_idx + 1}个目标"}
        
        target_config = targets[target_idx]
        target_base = Path(target_config.get('path', ''))
        
        # 初始化linker
        obfuscate_enabled = monitor.get('obfuscate_enabled', True)
        linker = FileLinker(obfuscate_enabled=obfuscate_enabled, db_engine=db_engine)
        
        # 临时更新映射表，让混淆器使用新名称
        if request.target_type == 'quark':
            old_name = mapping.quark_name
            mapping.quark_name = display_name
        elif request.target_type == 'baidu':
            old_name = mapping.baidu_name
            mapping.baidu_name = display_name
        elif request.target_type == 'xunlei':
            old_name = mapping.xunlei_name
            mapping.xunlei_name = display_name
        db.commit()
        
        success_count = 0
        failed_count = 0
        deleted_dirs = set()
        
        for record in records:
            try:
                source_file = Path(record.source_file)
                if not source_file.exists():
                    logger.warning(f"源文件不存在: {source_file}")
                    continue
                
                # 获取旧的目标文件路径
                if request.target_type == 'quark':
                    old_target_file = record.quark_target_file
                elif request.target_type == 'baidu':
                    old_target_file = record.baidu_target_file
                elif request.target_type == 'xunlei':
                    old_target_file = record.xunlei_target_file
                else:
                    old_target_file = None
                
                # 删除旧文件
                if old_target_file:
                    old_path = Path(old_target_file)
                    if old_path.exists():
                        old_path.unlink()
                        logger.info(f"已删除旧文件: {old_path}")
                        
                        # 收集需要删除的空目录
                        for parent in old_path.parents:
                            if parent.name.startswith('Season') or parent == old_path.parent.parent:
                                deleted_dirs.add(parent.parent)
                                break
                
                # 计算新的目标路径
                relative_path = source_file.relative_to(source_path)
                new_target_file = target_base / relative_path
                
                # 重新硬链接
                success, method, error, actual_target = linker.create_hardlink(
                    source_file, new_target_file,
                    source_base=source_path,
                    target_base=target_base
                )
                
                if success and actual_target:
                    # 更新记录
                    if request.target_type == 'quark':
                        record.quark_target_file = str(actual_target)
                        record.quark_synced_at = datetime.now()
                    elif request.target_type == 'baidu':
                        record.baidu_target_file = str(actual_target)
                        record.baidu_synced_at = datetime.now()
                    elif request.target_type == 'xunlei':
                        record.xunlei_target_file = str(actual_target)
                        record.xunlei_synced_at = datetime.now()
                    
                    record.updated_at = datetime.now()
                    success_count += 1
                    logger.info(f"重转成功: {source_file.name} -> {actual_target}")
                else:
                    failed_count += 1
                    logger.error(f"重转失败: {source_file.name}, 错误: {error}")
                    
            except Exception as e:
                failed_count += 1
                logger.error(f"处理文件失败 {record.source_file}: {e}")
        
        # 删除空目录
        for dir_path in deleted_dirs:
            if dir_path.exists():
                try:
                    shutil.rmtree(dir_path)
                    logger.info(f"已删除目录: {dir_path}")
                except Exception as e:
                    logger.error(f"删除目录失败 {dir_path}: {e}")
        
        db.commit()
        
        # 防失效处理：修改视频文件Hash
        if success_count > 0:
            try:
                # 从已重转的文件记录中提取实际的剧集根目录
                sample_record = db.query(LinkRecord).filter(
                    LinkRecord.original_name == request.original_name
                ).first()
                
                if sample_record:
                    # 根据target_type获取对应的target_file
                    target_file_str = None
                    if request.target_type == 'quark':
                        target_file_str = sample_record.quark_target_file
                    elif request.target_type == 'baidu':
                        target_file_str = sample_record.baidu_target_file
                    elif request.target_type == 'xunlei':
                        target_file_str = sample_record.xunlei_target_file
                    
                    if target_file_str:
                        # 例如: /media/夸克网盘/剧集/国产剧集/C钞钞感贡米(2025)/Season 1/S01E01.mkv
                        # 提取剧集根目录: /media/夸克网盘/剧集/国产剧集/C钞钞感贡米(2025)
                        target_file_path = Path(target_file_str)
                        
                        # 向上找2级：文件 -> Season 1 -> 剧集根目录
                        if target_file_path.exists():
                            target_show_dir = target_file_path.parent.parent
                            
                            logger.info(f"开始防失效处理: {target_show_dir}")
                            processed_count = process_anti_ban(target_show_dir)
                            logger.info(f"防失效处理完成: 处理了 {processed_count} 个视频文件")
                        else:
                            logger.warning(f"目标文件不存在: {target_file_path}")
                    else:
                        logger.warning(f"未找到目标文件路径")
                else:
                    logger.warning(f"未找到重转记录")
            except Exception as e:
                logger.error(f"防失效处理失败: {e}")
                import traceback
                traceback.print_exc()
        
        # 触发TaoSync同步
        if success_count > 0:
            try:
                from backend.main import monitor_service
                if monitor_service and monitor_service.handlers:
                    for handler in monitor_service.handlers:
                        if hasattr(handler, 'taosync_queue') and handler.taosync_queue:
                            logger.info(f"重转完成，触发TaoSync同步（{success_count}个文件）")
                            handler.taosync_queue.trigger_now(file_count=success_count)
                            break
            except Exception as e:
                logger.error(f"触发TaoSync失败: {e}")
        
        target_names = {'quark': '夸克', 'baidu': '百度', 'xunlei': '迅雷'}
        target_name = target_names.get(request.target_type, '未知')
        return {
            "success": True,
            "message": f"重转完成: 成功 {success_count} 个文件，失败 {failed_count} 个",
            "data": {
                "success_count": success_count,
                "failed_count": failed_count,
                "target_type": target_name,
                "display_name": display_name
            }
        }
        
    except Exception as e:
        logger.error(f"重转失败: {e}")
        db.rollback()
        return {
            "success": False,
            "message": f"重转失败: {str(e)}"
        }
