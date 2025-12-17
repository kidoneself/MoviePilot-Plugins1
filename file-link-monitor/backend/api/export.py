from fastapi import APIRouter, Depends, Response
from sqlalchemy.orm import Session
from pathlib import Path
import logging
from io import BytesIO
from datetime import datetime

from backend.models import LinkRecord, get_session

router = APIRouter()
logger = logging.getLogger(__name__)


def get_db():
    """依赖注入：获取数据库会话"""
    from backend.main import db_engine
    session = get_session(db_engine)
    try:
        yield session
    finally:
        session.close()


@router.get("/export/name-mapping")
async def export_name_mapping(db: Session = Depends(get_db)):
    """
    导出原名→新名的Excel映射表
    
    Returns:
        Excel文件（二进制流）
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # 查询所有成功的硬链接记录
        records = db.query(LinkRecord).filter(
            LinkRecord.status == "success"
        ).all()
        
        # 提取文件夹名称映射（去重）
        mapping = {}
        for record in records:
            source_path = Path(record.source_file)
            target_path = Path(record.target_file)
            
            # 获取剧集名称（去掉Season和文件名）
            source_parts = source_path.parts
            target_parts = target_path.parts
            
            # 找到分类文件夹后的第一个文件夹（剧集名）
            # 路径格式：.../分类/二级分类/剧名/Season X/文件
            if len(source_parts) >= 4 and len(target_parts) >= 4:
                # 从后往前数：文件名、Season、剧名、二级分类、分类
                source_show = source_parts[-3] if 'Season' in source_parts[-2] else source_parts[-2]
                target_show = target_parts[-3] if 'Season' in target_parts[-2] else target_parts[-2]
                
                # 去重存储
                if source_show not in mapping:
                    mapping[source_show] = target_show
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "名称映射表"
        
        # 设置表头
        headers = ["原名称", "混淆后名称", "说明"]
        ws.append(headers)
        
        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 填充数据
        row_num = 2
        for source_name in sorted(mapping.keys()):
            target_name = mapping[source_name]
            
            # 判断是否混淆（有变化）
            if source_name == target_name:
                note = "未混淆（分类文件夹或英文名）"
            else:
                note = "已混淆"
            
            ws.append([source_name, target_name, note])
            
            # 设置对齐
            for col_num in range(1, 4):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            
            row_num += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 30
        
        # 添加统计信息
        ws.append([])
        ws.append([f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.append([f"总计: {len(mapping)} 个剧集"])
        ws.append([f"已混淆: {sum(1 for s, t in mapping.items() if s != t)} 个"])
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        filename = f"name_mapping_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        logger.info(f"✅ 成功导出名称映射表: {len(mapping)} 条记录")
        
        # 返回Excel文件
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except ImportError:
        logger.error("openpyxl库未安装，无法导出Excel")
        return {
            "success": False,
            "message": "openpyxl库未安装，请安装: pip install openpyxl"
        }
    except Exception as e:
        logger.error(f"导出失败: {e}")
        return {
            "success": False,
            "message": f"导出失败: {str(e)}"
        }
