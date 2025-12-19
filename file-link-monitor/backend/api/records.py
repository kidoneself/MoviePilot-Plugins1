from fastapi import APIRouter, Depends, Response
from sqlalchemy.orm import Session
from sqlalchemy import func, desc
from typing import Optional
import logging
from io import BytesIO
from datetime import datetime
from pathlib import Path

from backend.models import LinkRecord, get_session

router = APIRouter()
logger = logging.getLogger(__name__)


def get_db():
    """依赖注入：获取数据库会话"""
    from backend.main import db_engine
    session = get_session(db_engine)
    try:
        yield session
    finally:
        session.close()


@router.get("/records")
async def get_records(
    page: int = 1,
    page_size: int = 50,
    status: Optional[str] = None,
    group_by: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """获取硬链接记录"""
    try:
        from pathlib import Path
        
        query = db.query(LinkRecord)
        
        # 搜索源文件路径
        if search:
            query = query.filter(LinkRecord.source_file.like(f'%{search}%'))
        
        # 状态筛选（新表结构不支持status）
        # if status:
        #     query = query.filter(LinkRecord.status == status)
        
        # 总数
        total = query.count()
        
        # 如果是分组模式，先获取所有记录再分组，最后对分组结果分页
        if group_by:
            # 获取所有记录（不分页）
            # 按源文件路径排序（这样集数会自然排序）
            records = query.order_by(LinkRecord.source_file.asc()).all()
        else:
            # 不分组时，按创建时间倒序
            records = query.order_by(desc(LinkRecord.created_at))\
                          .offset((page - 1) * page_size)\
                          .limit(page_size)\
                          .all()
        
        # 按源目录分组
        if group_by == "source_dir":
            from collections import defaultdict
            import re
            
            groups = defaultdict(list)
            
            for record in records:
                # 获取源文件路径，尝试找到剧名目录（包含年份的目录）
                source_path = Path(record.source_file)
                show_name = None
                
                # 从路径中找包含年份的目录作为剧名
                for part in source_path.parts:
                    if re.search(r'\(\d{4}\)', part):
                        show_name = part
                        break
                
                # 如果没找到年份目录，用父目录名
                if not show_name:
                    show_name = source_path.parent.name if source_path.parent.name else "根目录"
                
                groups[show_name].append({
                    "id": record.id,
                    "source_file": record.source_file,
                    "quark_target_file": record.quark_target_file,
                    "baidu_target_file": record.baidu_target_file,
                    "file_size": record.file_size,
                    "created_at": record.created_at.strftime("%Y-%m-%d %H:%M:%S") if record.created_at else None
                })
            
            # 转换为列表格式
            grouped_data = []
            for dir_name, items in groups.items():
                grouped_data.append({
                    "dir_name": dir_name,
                    "count": len(items),
                    "records": items
                })
            
            # 对分组结果分页
            total_groups = len(grouped_data)
            start_idx = (page - 1) * page_size
            end_idx = start_idx + page_size
            paginated_groups = grouped_data[start_idx:end_idx]
            
            return {
                "success": True,
                "total": total_groups,
                "page": page,
                "page_size": page_size,
                "grouped": True,
                "data": paginated_groups
            }
        elif group_by == "source_file":
            from collections import defaultdict
            
            groups = defaultdict(list)
            
            for record in records:
                # 按源文件完整路径分组
                source_file = record.source_file
                
                groups[source_file].append({
                    "id": record.id,
                    "source_file": record.source_file,
                    "quark_target_file": record.quark_target_file,
                    "baidu_target_file": record.baidu_target_file,
                    "file_size": record.file_size,
                    "created_at": record.created_at.strftime("%Y-%m-%d %H:%M:%S") if record.created_at else None
                })
            
            # 转换为列表格式
            grouped_data = []
            for source_file, items in groups.items():
                # 获取文件名作为显示名称
                file_name = Path(source_file).name
                grouped_data.append({
                    "dir_name": file_name,  # 使用文件名作为组名
                    "source_file": source_file,  # 完整路径
                    "count": len(items),
                    "records": items
                })
            
            # 对分组结果分页
            total_groups = len(grouped_data)
            start_idx = (page - 1) * page_size
            end_idx = start_idx + page_size
            paginated_groups = grouped_data[start_idx:end_idx]
            
            return {
                "success": True,
                "grouped": True,
                "data": paginated_groups,
                "total": total_groups,
                "page": page,
                "page_size": page_size
            }
        elif group_by == "target_show":
            # 按目标目录+剧集名分组统计
            from collections import defaultdict
            import re
            
            # 目标目录 -> 剧集名 -> 记录列表
            target_shows = defaultdict(lambda: defaultdict(list))
            
            for record in records:
                # 优先使用quark_target_file，如果没有则使用baidu_target_file
                target_file = record.quark_target_file or record.baidu_target_file
                if not target_file:
                    continue
                target_path = Path(target_file)
                # 提取目标根目录（如 /media/网盘测试1）
                target_base = str(target_path.parts[0:3] if len(target_path.parts) >= 3 else target_path.parts[0:2])
                
                # 尝试从路径中提取剧集名（包含年份的目录）
                show_name = None
                for part in target_path.parts:
                    # 匹配包含年份的目录，如 "大生意人 (2025)"
                    if re.search(r'\(\d{4}\)', part):
                        show_name = part
                        break
                
                if not show_name:
                    show_name = "其他文件"
                
                target_shows[target_base][show_name].append(record)
            
            # 组装数据
            grouped_data = []
            for target_base, shows in target_shows.items():
                target_records = []
                for show_name, records_list in shows.items():
                    # 统计集数和大小
                    total_size = sum(r.file_size for r in records_list)
                    count = len(records_list)
                    
                    target_records.append({
                        "show_name": show_name,
                        "count": count,
                        "total_size": total_size,
                        "records": [{
                            "id": r.id,
                            "source_file": r.source_file,
                            "quark_target_file": r.quark_target_file,
                            "baidu_target_file": r.baidu_target_file,
                            "file_size": r.file_size,
                            "created_at": r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else None
                        } for r in records_list]
                    })
                
                grouped_data.append({
                    "dir_name": Path(target_base).name if target_base else "未知目录",
                    "target_base": target_base,
                    "count": sum(len(r["records"]) for r in target_records),
                    "shows": target_records
                })
            
            # 对网盘级别分页（一页显示几个网盘）
            total_groups = len(grouped_data)
            start_idx = (page - 1) * page_size
            end_idx = start_idx + page_size
            paginated_groups = grouped_data[start_idx:end_idx]
            
            return {
                "success": True,
                "grouped": True,
                "group_type": "target_show",
                "data": paginated_groups,
                "total": total_groups,
                "page": page,
                "page_size": page_size
            }
        
        # 不分组，返回原格式
        data = []
        for record in records:
            data.append({
                "id": record.id,
                "source_file": record.source_file,
                "original_name": record.original_name,
                "quark_target_file": record.quark_target_file,
                "baidu_target_file": record.baidu_target_file,
                "xunlei_target_file": record.xunlei_target_file,
                "file_size": record.file_size,
                "created_at": record.created_at.strftime("%Y-%m-%d %H:%M:%S") if record.created_at else None
            })
        
        return {
            "success": True,
            "total": total,
            "page": page,
            "page_size": page_size,
            "grouped": False,
            "data": data
        }
    except Exception as e:
        logger.error(f"获取记录失败: {e}")
        return {"success": False, "message": str(e)}


@router.get("/stats")
async def get_stats(db: Session = Depends(get_db)):
    """获取统计信息"""
    try:
        # 总记录数
        total_records = db.query(func.count(LinkRecord.id)).scalar()
        
        # 同步统计（有任一网盘同步即算成功）
        success_count = db.query(func.count(LinkRecord.id))\
                         .filter((LinkRecord.quark_target_file.isnot(None)) | (LinkRecord.baidu_target_file.isnot(None))).scalar()
        failed_count = 0  # 新表结构不记录失败
        
        # 总大小
        total_size = db.query(func.sum(LinkRecord.file_size)).scalar() or 0
        
        # 今天的记录
        from datetime import datetime, timedelta
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        today_count = db.query(func.count(LinkRecord.id))\
                       .filter(LinkRecord.created_at >= today).scalar()
        
        # 最近10条记录
        recent = db.query(LinkRecord)\
                  .order_by(desc(LinkRecord.created_at))\
                  .limit(10)\
                  .all()
        
        recent_data = []
        for r in recent:
            recent_data.append({
                "source_file": r.source_file,
                "original_name": r.original_name,
                "quark_target_file": r.quark_target_file,
                "baidu_target_file": r.baidu_target_file,
                "file_size": r.file_size,
                "created_at": r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else None
            })
        
        return {
            "success": True,
            "data": {
                "total_records": total_records,
                "success_count": success_count,
                "failed_count": failed_count,
                "total_size": total_size,
                "today_count": today_count,
                "recent": recent_data
            }
        }
    except Exception as e:
        logger.error(f"获取统计信息失败: {e}")
        return {"success": False, "message": str(e)}


@router.get("/today-sync")
async def get_today_sync(db: Session = Depends(get_db)):
    """获取今日同步明细，按网盘和目录结构分组"""
    try:
        from datetime import datetime, timedelta
        import re
        from collections import defaultdict
        
        # 获取今天的记录
        today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        records = db.query(LinkRecord).filter(
            LinkRecord.created_at >= today
        ).all()
        
        logger.info(f"今日同步统计：查询到 {len(records)} 条记录")
        
        # 按网盘分组
        result = {
            'quark': defaultdict(lambda: defaultdict(lambda: defaultdict(list))),
            'baidu': defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
        }
        
        # 定义分类列表
        category1_list = ['剧集', '电影', '动漫', '其他']
        category2_list = [
            '港台电影', '国产电影', '日韩电影', '欧美电影', '动画电影',
            '港台剧集', '国产剧集', '日韩剧集', '南亚剧集', '欧美剧集',
            '国产动漫', '欧美动漫', '日本番剧',
            '纪录影片', '综艺节目'
        ]
        
        for record in records:
            source_path = Path(record.source_file)
            parts = list(source_path.parts)
            filename = source_path.name
            show_name = record.original_name
            
            if not show_name:
                logger.debug(f"跳过：无剧名 {record.source_file}")
                continue
            
            # 从路径中找分类
            category1 = ''
            category2 = ''
            
            for part in parts:
                if part in category1_list:
                    category1 = part
                elif part in category2_list:
                    category2 = part
            
            if not category1 or not category2:
                logger.debug(f"跳过：分类不全 {record.source_file}, cat1={category1}, cat2={category2}")
                continue
            
            logger.debug(f"解析成功：{category1} > {category2} > {show_name} > {filename}")
            
            # 夸克网盘
            if record.quark_target_file:
                result['quark'][category1][category2][show_name].append(filename)
            
            # 百度网盘
            if record.baidu_target_file:
                result['baidu'][category1][category2][show_name].append(filename)
        
        # 转换为普通dict便于JSON序列化
        output = {
            'quark': {k: {k2: {k3: sorted(v3) for k3, v3 in v2.items()} for k2, v2 in v.items()} for k, v in result['quark'].items()},
            'baidu': {k: {k2: {k3: sorted(v3) for k3, v3 in v2.items()} for k2, v2 in v.items()} for k, v in result['baidu'].items()}
        }
        
        return {
            "success": True,
            "data": output
        }
        
    except Exception as e:
        logger.error(f"获取今日同步统计失败: {e}")
        return {"success": False, "message": str(e)}


@router.delete("/records/batch")
async def batch_delete_records(search: str, db: Session = Depends(get_db)):
    """批量删除记录（根据搜索条件）"""
    try:
        if not search:
            return {"success": False, "message": "搜索条件不能为空"}
        
        # 查找所有匹配的记录
        query = db.query(LinkRecord).filter(LinkRecord.source_file.like(f'%{search}%'))
        count = query.count()
        
        if count == 0:
            return {"success": False, "message": "未找到匹配的记录"}
        
        # 删除所有匹配的记录
        query.delete(synchronize_session=False)
        db.commit()
        
        logger.info(f"批量删除了 {count} 条记录（搜索词: {search}）")
        return {"success": True, "message": f"已删除 {count} 条记录", "count": count}
    except Exception as e:
        logger.error(f"批量删除记录失败: {e}")
        db.rollback()
        return {"success": False, "message": str(e)}


@router.delete("/records/by-show")
async def delete_records_by_show(
    show_name: str,
    db: Session = Depends(get_db)
):
    """
    删除指定剧集的所有硬链接记录，并删除目标目录中的对应文件夹
    
    Args:
        show_name: 剧集名称（原始名称，会匹配路径中包含该名称的所有记录）
    """
    try:
        from pathlib import Path
        import shutil
        
        # 查询包含该剧集名称的所有记录
        records = db.query(LinkRecord).filter(
            LinkRecord.source_file.like(f'%{show_name}%')
        ).all()
        
        if not records:
            return {
                "success": False,
                "message": f"未找到包含 '{show_name}' 的记录"
            }
        
        count = len(records)
        
        # 收集需要删除的目录（去重）
        dirs_to_delete = set()
        for record in records:
            # 删除夸克和百度的目录
            for target_file in [record.quark_target_file, record.baidu_target_file]:
                if not target_file:
                    continue
                target_path = Path(target_file)
                logger.info(f"处理记录: {target_file}")
                
                # 找到剧集目录（包含Season的父目录）
                show_dir = None
                for parent in target_path.parents:
                    if 'Season' in parent.name or 'season' in parent.name:
                        show_dir = parent.parent
                        logger.info(f"  找到Season目录，剧集目录: {show_dir}")
                        break
                
                # 如果没找到Season目录，默认向上2级
                if not show_dir:
                    show_dir = target_path.parent.parent
                    logger.info(f"  未找到Season目录，使用默认向上2级: {show_dir}")
                
                if show_dir:
                    logger.info(f"  目录存在检查: {show_dir.exists()} - {show_dir}")
                    if show_dir.exists():
                        dirs_to_delete.add(str(show_dir))
                        logger.info(f"  ✓ 添加到删除列表: {show_dir}")
                    else:
                        logger.warning(f"  ✗ 目录不存在: {show_dir}")
        
        # 删除所有匹配的记录
        for record in records:
            db.delete(record)
        
        db.commit()
        
        # 删除目录
        deleted_dirs = []
        for dir_path in dirs_to_delete:
            try:
                if Path(dir_path).exists():
                    shutil.rmtree(dir_path)
                    deleted_dirs.append(dir_path)
                    logger.info(f"已删除目录: {dir_path}")
            except Exception as e:
                logger.error(f"删除目录失败 {dir_path}: {e}")
        
        return {
            "success": True,
            "message": f"已删除 {count} 条记录和 {len(deleted_dirs)} 个目录",
            "count": count,
            "deleted_dirs": deleted_dirs
        }
    except Exception as e:
        logger.error(f"删除剧集记录失败: {e}")
        db.rollback()
        return {"success": False, "message": str(e)}


@router.delete("/records/{record_id}")
async def delete_record(record_id: int, db: Session = Depends(get_db)):
    """删除记录"""
    try:
        record = db.query(LinkRecord).filter(LinkRecord.id == record_id).first()
        if not record:
            return {"success": False, "message": "记录不存在"}
        
        db.delete(record)
        db.commit()
        
        return {"success": True, "message": "删除成功"}
    except Exception as e:
        logger.error(f"删除记录失败: {e}")
        return {"success": False, "message": str(e)}


@router.post("/records/{record_id}/retry")
async def retry_link(record_id: int, db: Session = Depends(get_db)):
    """重试硬链接"""
    try:
        from pathlib import Path
        from backend.utils.linker import FileLinker
        from datetime import datetime
        import yaml
        
        # 获取记录
        record = db.query(LinkRecord).filter(LinkRecord.id == record_id).first()
        if not record:
            return {"success": False, "message": "记录不存在"}
        
        source_file = Path(record.source_file)
        target_file = Path(record.target_file)
        
        # 检查源文件是否存在
        if not source_file.exists():
            return {"success": False, "message": "源文件不存在"}
        
        # 从配置文件读取混淆设置
        config_path = Path(__file__).parent.parent.parent / "config.yaml"
        with open(config_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
        
        # 查找该源文件所属的监控配置
        obfuscate_enabled = False
        for monitor in config.get('monitors', []):
            source_dir = Path(monitor['source'])
            if source_file.is_relative_to(source_dir):
                obfuscate_enabled = monitor.get('obfuscate_enabled', False)
                break
        
        # 重新创建硬链接（使用混淆设置）
        linker = FileLinker(obfuscate_enabled=obfuscate_enabled)
        success, method, error = linker.create_hardlink(source_file, target_file)
        
        # 更新记录
        record.link_method = method
        record.status = "success" if success else "failed"
        record.error_msg = error
        record.created_at = datetime.now()
        
        db.commit()
        
        if success:
            return {
                "success": True,
                "message": f"重试成功，使用{method}方式"
            }
        else:
            return {
                "success": False,
                "message": f"重试失败：{error}"
            }
    except Exception as e:
        logger.error(f"重试硬链接失败: {e}")
        db.rollback()
        return {"success": False, "message": str(e)}


@router.post("/records/{record_id}/resync")
async def resync_link(record_id: int, db: Session = Depends(get_db)):
    """重新同步（删除记录并重新创建硬链接）"""
    try:
        from pathlib import Path
        from backend.utils.linker import FileLinker
        from datetime import datetime
        import yaml
        
        # 获取记录
        record = db.query(LinkRecord).filter(LinkRecord.id == record_id).first()
        if not record:
            return {"success": False, "message": "记录不存在"}
        
        source_file = Path(record.source_file)
        target_file = Path(record.target_file)
        
        # 检查源文件是否存在
        if not source_file.exists():
            return {"success": False, "message": "源文件不存在"}
        
        # 从配置文件读取混淆设置
        config_path = Path(__file__).parent.parent.parent / "config.yaml"
        with open(config_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
        
        # 查找该源文件所属的监控配置
        obfuscate_enabled = False
        for monitor in config.get('monitors', []):
            source_dir = Path(monitor['source'])
            if source_file.is_relative_to(source_dir):
                obfuscate_enabled = monitor.get('obfuscate_enabled', False)
                break
        
        # 删除旧记录
        db.delete(record)
        db.commit()
        
        # 重新创建硬链接（使用混淆设置）
        linker = FileLinker(obfuscate_enabled=obfuscate_enabled)
        success, method, error = linker.create_hardlink(source_file, target_file)
        
        # 创建新记录
        new_record = LinkRecord(
            source_file=str(source_file),
            target_file=str(target_file),
            file_size=source_file.stat().st_size,
            link_method=method,
            status="success" if success else "failed",
            error_msg=error
        )
        db.add(new_record)
        db.commit()
        
        if success:
            return {
                "success": True,
                "message": f"重新同步成功，使用{method}方式"
            }
        else:
            return {
                "success": False,
                "message": f"重新同步失败：{error}"
            }
    except Exception as e:
        logger.error(f"重新同步失败: {e}")
        db.rollback()
        return {"success": False, "message": str(e)}


@router.get("/export/records")
async def export_records(
    status: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    导出链接记录为Excel文件
    
    Args:
        status: 过滤状态
        search: 搜索关键词
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from pathlib import Path
        
        query = db.query(LinkRecord)
        
        # if status:
        #     query = query.filter(LinkRecord.status == status)
        
        if search:
            query = query.filter(LinkRecord.source_file.like(f'%{search}%'))
        
        records = query.order_by(LinkRecord.created_at.desc()).all()
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "链接记录"
        
        # 设置表头
        headers = ["源文件", "夸克目标", "百度目标", "文件大小(MB)", "创建时间"]
        ws.append(headers)
        
        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 填充数据
        row_num = 2
        for record in records:
            file_size_mb = round(record.file_size / (1024 * 1024), 2) if record.file_size else 0
            created_at = record.created_at.strftime('%Y-%m-%d %H:%M:%S') if record.created_at else ''
            
            ws.append([
                record.source_file,
                record.quark_target_file or '-',
                record.baidu_target_file or '-',
                file_size_mb,
                created_at
            ])
            
            # 设置对齐
            for col_num in range(1, 6):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            
            row_num += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 60
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 40
        
        # 添加统计信息
        ws.append([])
        ws.append([f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.append([f"总计: {len(records)} 条记录"])
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名
        filename = f"link_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        logger.info(f"✅ 成功导出链接记录: {len(records)} 条")
        
        # 返回Excel文件
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except ImportError:
        logger.error("openpyxl库未安装，无法导出Excel")
        return {
            "success": False,
            "message": "openpyxl库未安装，请安装: pip install openpyxl"
        }
    except Exception as e:
        logger.error(f"导出失败: {e}")
        return {
            "success": False,
            "message": f"导出失败: {str(e)}"
        }
