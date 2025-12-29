"""
è‡ªå®šä¹‰åç§°æ˜ å°„ç®¡ç†API
"""
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from fastapi.responses import Response
from sqlalchemy.orm import Session
from sqlalchemy import func, or_, and_, distinct
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime
from io import BytesIO
import logging
import shutil
import subprocess
import uuid
import os
from pathlib import Path

from backend.models import CustomNameMapping, LinkRecord, get_session, get_db
from backend.utils.linker import FileLinker
from backend.utils.obfuscator import FolderObfuscator

router = APIRouter()
logger = logging.getLogger(__name__)



class MappingCreate(BaseModel):
    """åˆ›å»ºæ˜ å°„è¯·æ±‚"""
    original_name: str
    category: Optional[str] = None
    quark_name: Optional[str] = None
    baidu_name: Optional[str] = None
    xunlei_name: Optional[str] = None
    note: Optional[str] = None
    baidu_link: Optional[str] = None
    quark_link: Optional[str] = None
    xunlei_link: Optional[str] = None
    sync_to_quark: Optional[bool] = True
    sync_to_baidu: Optional[bool] = True
    sync_to_xunlei: Optional[bool] = True


class MappingUpdate(BaseModel):
    """æ›´æ–°æ˜ å°„è¯·æ±‚"""
    category: Optional[str] = None
    quark_name: Optional[str] = None
    baidu_name: Optional[str] = None
    xunlei_name: Optional[str] = None
    enabled: Optional[bool] = None
    is_completed: Optional[bool] = None
    note: Optional[str] = None
    baidu_link: Optional[str] = None
    quark_link: Optional[str] = None
    xunlei_link: Optional[str] = None
    sync_to_quark: Optional[bool] = None
    sync_to_baidu: Optional[bool] = None
    sync_to_xunlei: Optional[bool] = None


@router.get("/mappings")
async def get_mappings(
    page: int = 1,
    page_size: int = 20,
    enabled: Optional[bool] = None,
    search: Optional[str] = None,
    media_type: Optional[str] = None,
    is_completed: Optional[bool] = None,
    category: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    è·å–æ‰€æœ‰è‡ªå®šä¹‰åç§°æ˜ å°„ï¼ˆåˆ†é¡µï¼‰
    
    Args:
        page: é¡µç ï¼ˆä»1å¼€å§‹ï¼‰
        page_size: æ¯é¡µæ•°é‡
        enabled: è¿‡æ»¤å¯ç”¨çŠ¶æ€
        search: æœç´¢åŸåæˆ–è‡ªå®šä¹‰å
        media_type: åª’ä½“ç±»å‹ç­›é€‰ (movie/tv)
        is_completed: å®Œç»“çŠ¶æ€ç­›é€‰ (True/False)
        category: äºŒçº§åˆ†ç±»ç­›é€‰ (å¦‚: "ç”µå½±/å›½äº§ç”µå½±")
    """
    try:
        query = db.query(CustomNameMapping)
        
        if enabled is not None:
            query = query.filter(CustomNameMapping.enabled == enabled)
        
        if search:
            query = query.filter(
                (CustomNameMapping.original_name.like(f'%{search}%')) |
                (CustomNameMapping.quark_name.like(f'%{search}%')) |
                (CustomNameMapping.baidu_name.like(f'%{search}%')) |
                (CustomNameMapping.xunlei_name.like(f'%{search}%'))
            )
        
        # åª’ä½“ç±»å‹ç­›é€‰
        if media_type:
            query = query.filter(CustomNameMapping.media_type == media_type)
        
        # å®Œç»“çŠ¶æ€ç­›é€‰
        if is_completed is not None:
            query = query.filter(CustomNameMapping.is_completed == is_completed)
        
        # äºŒçº§åˆ†ç±»ç­›é€‰
        if category:
            query = query.filter(CustomNameMapping.category == category)
        
        # æ€»æ•°
        total = query.count()
        
        # åˆ†é¡µ
        mappings = query.order_by(CustomNameMapping.created_at.desc())\
            .offset((page - 1) * page_size)\
            .limit(page_size)\
            .all()
        
        return {
            "success": True,
            "data": [
                {
                    "id": m.id,
                    "original_name": m.original_name,
                    "category": m.category,
                    "quark_name": m.quark_name,
                    "baidu_name": m.baidu_name,
                    "xunlei_name": m.xunlei_name,
                    "enabled": m.enabled,
                    "is_completed": m.is_completed,
                    "note": m.note,
                    "baidu_link": m.baidu_link,
                    "quark_link": m.quark_link,
                    "xunlei_link": m.xunlei_link,
                    "sync_to_quark": m.sync_to_quark if hasattr(m, 'sync_to_quark') else True,
                    "sync_to_baidu": m.sync_to_baidu if hasattr(m, 'sync_to_baidu') else True,
                    "sync_to_xunlei": m.sync_to_xunlei if hasattr(m, 'sync_to_xunlei') else True,
                    # TMDb å…ƒæ•°æ®
                    "tmdb_id": m.tmdb_id if hasattr(m, 'tmdb_id') else None,
                    "poster_url": m.poster_url if hasattr(m, 'poster_url') else None,
                    "overview": m.overview if hasattr(m, 'overview') else None,
                    "media_type": m.media_type if hasattr(m, 'media_type') else None,
                    "created_at": m.created_at.isoformat() if m.created_at else None,
                    "updated_at": m.updated_at.isoformat() if m.updated_at else None
                }
                for m in mappings
            ],
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (total + page_size - 1) // page_size
        }
    except Exception as e:
        logger.error(f"è·å–æ˜ å°„å¤±è´¥: {e}")
        return {"success": False, "message": str(e)}


@router.post("/mappings")
async def create_mapping(mapping: MappingCreate, db: Session = Depends(get_db)):
    """
    åˆ›å»ºè‡ªå®šä¹‰åç§°æ˜ å°„
    
    Args:
        mapping: æ˜ å°„æ•°æ®
    """
    try:
        # æ£€æŸ¥æ˜¯å¦å·²å­˜åœ¨
        existing = db.query(CustomNameMapping).filter(
            CustomNameMapping.original_name == mapping.original_name
        ).first()
        
        if existing:
            return {
                "success": False,
                "message": f"æ˜ å°„å·²å­˜åœ¨: {mapping.original_name}"
            }
        
        # åˆ›å»ºæ˜ å°„ï¼ˆè‡ªåŠ¨å»é™¤å‰åç©ºæ ¼ï¼‰
        new_mapping = CustomNameMapping(
            original_name=mapping.original_name,
            category=mapping.category,
            quark_name=mapping.quark_name.strip() if mapping.quark_name else None,
            baidu_name=mapping.baidu_name.strip() if mapping.baidu_name else None,
            xunlei_name=mapping.xunlei_name.strip() if mapping.xunlei_name else None,
            note=mapping.note,
            baidu_link=mapping.baidu_link,
            quark_link=mapping.quark_link,
            xunlei_link=mapping.xunlei_link,
            sync_to_quark=mapping.sync_to_quark,
            sync_to_baidu=mapping.sync_to_baidu,
            sync_to_xunlei=mapping.sync_to_xunlei
        )
        db.add(new_mapping)
        db.commit()
        db.refresh(new_mapping)
        
        logger.info(f"âœ… åˆ›å»ºæ˜ å°„: {mapping.original_name} (å¤¸å…‹:{mapping.quark_name}, ç™¾åº¦:{mapping.baidu_name})")
        
        return {
            "success": True,
            "message": "æ˜ å°„åˆ›å»ºæˆåŠŸ",
            "data": {
                "id": new_mapping.id,
                "original_name": new_mapping.original_name,
                "category": new_mapping.category,
                "quark_name": new_mapping.quark_name,
                "baidu_name": new_mapping.baidu_name,
                "xunlei_name": new_mapping.xunlei_name,
                "enabled": new_mapping.enabled,
                "note": new_mapping.note
            }
        }
    except Exception as e:
        db.rollback()
        logger.error(f"åˆ›å»ºæ˜ å°„å¤±è´¥: {e}")
        return {"success": False, "message": str(e)}


@router.get("/mappings/{mapping_id}")
async def get_mapping(
    mapping_id: int,
    db: Session = Depends(get_db)
):
    """
    è·å–å•ä¸ªæ˜ å°„è¯¦æƒ…
    
    Args:
        mapping_id: æ˜ å°„ID
    """
    try:
        mapping = db.query(CustomNameMapping).filter(
            CustomNameMapping.id == mapping_id
        ).first()
        
        if not mapping:
            return {"success": False, "message": "æ˜ å°„ä¸å­˜åœ¨"}
        
        return {
            "success": True,
            "data": {
                "id": mapping.id,
                "original_name": mapping.original_name,
                "category": mapping.category,
                "quark_name": mapping.quark_name,
                "baidu_name": mapping.baidu_name,
                "xunlei_name": mapping.xunlei_name,
                "enabled": mapping.enabled,
                "is_completed": mapping.is_completed,
                "created_at": mapping.created_at.isoformat() if mapping.created_at else None,
                "updated_at": mapping.updated_at.isoformat() if mapping.updated_at else None,
                "note": mapping.note,
                "baidu_link": mapping.baidu_link,
                "quark_link": mapping.quark_link,
                "xunlei_link": mapping.xunlei_link,
                "sync_to_quark": mapping.sync_to_quark,
                "sync_to_baidu": mapping.sync_to_baidu,
                "sync_to_xunlei": mapping.sync_to_xunlei,
                "tmdb_id": mapping.tmdb_id,
                "media_type": mapping.media_type,
                "poster_url": mapping.poster_url,
                "overview": mapping.overview
            }
        }
    except Exception as e:
        logger.error(f"è·å–æ˜ å°„å¤±è´¥: {e}")
        return {"success": False, "message": str(e)}


@router.put("/mappings/{mapping_id}")
async def update_mapping(
    mapping_id: int,
    mapping: MappingUpdate,
    db: Session = Depends(get_db)
):
    """
    æ›´æ–°è‡ªå®šä¹‰åç§°æ˜ å°„
    
    Args:
        mapping_id: æ˜ å°„ID
        mapping: æ›´æ–°æ•°æ®
    """
    try:
        existing = db.query(CustomNameMapping).filter(
            CustomNameMapping.id == mapping_id
        ).first()
        
        if not existing:
            return {"success": False, "message": "æ˜ å°„ä¸å­˜åœ¨"}
        
        # æ›´æ–°å­—æ®µï¼ˆè‡ªåŠ¨å»é™¤å‰åç©ºæ ¼ï¼‰
        if mapping.category is not None:
            existing.category = mapping.category
        if mapping.quark_name is not None:
            existing.quark_name = mapping.quark_name.strip() if mapping.quark_name else None
        if mapping.baidu_name is not None:
            existing.baidu_name = mapping.baidu_name.strip() if mapping.baidu_name else None
        if mapping.xunlei_name is not None:
            existing.xunlei_name = mapping.xunlei_name.strip() if mapping.xunlei_name else None
        if mapping.enabled is not None:
            existing.enabled = mapping.enabled
        if mapping.is_completed is not None:
            existing.is_completed = mapping.is_completed
        if mapping.note is not None:
            existing.note = mapping.note
        if mapping.baidu_link is not None:
            existing.baidu_link = mapping.baidu_link
        if mapping.quark_link is not None:
            existing.quark_link = mapping.quark_link
        if mapping.xunlei_link is not None:
            existing.xunlei_link = mapping.xunlei_link
        if mapping.sync_to_quark is not None:
            existing.sync_to_quark = mapping.sync_to_quark
        if mapping.sync_to_baidu is not None:
            existing.sync_to_baidu = mapping.sync_to_baidu
        if mapping.sync_to_xunlei is not None:
            existing.sync_to_xunlei = mapping.sync_to_xunlei
        
        db.commit()
        db.refresh(existing)
        
        logger.info(f"âœ… æ›´æ–°æ˜ å°„: {existing.original_name} (å¤¸å…‹:{existing.quark_name}, ç™¾åº¦:{existing.baidu_name})")
        
        return {
            "success": True,
            "message": "æ˜ å°„æ›´æ–°æˆåŠŸ",
            "data": {
                "id": existing.id,
                "original_name": existing.original_name,
                "category": existing.category,
                "quark_name": existing.quark_name,
                "baidu_name": existing.baidu_name,
                "xunlei_name": existing.xunlei_name,
                "enabled": existing.enabled,
                "note": existing.note
            }
        }
    except Exception as e:
        db.rollback()
        logger.error(f"æ›´æ–°æ˜ å°„å¤±è´¥: {e}")
        return {"success": False, "message": str(e)}


@router.delete("/mappings/{mapping_id}")
async def delete_mapping(mapping_id: int, db: Session = Depends(get_db)):
    """
    åˆ é™¤è‡ªå®šä¹‰åç§°æ˜ å°„
    
    Args:
        mapping_id: æ˜ å°„ID
    """
    try:
        mapping = db.query(CustomNameMapping).filter(
            CustomNameMapping.id == mapping_id
        ).first()
        
        if not mapping:
            return {"success": False, "message": "æ˜ å°„ä¸å­˜åœ¨"}
        
        original_name = mapping.original_name
        db.delete(mapping)
        db.commit()
        
        logger.info(f"âœ… åˆ é™¤æ˜ å°„: {original_name}")
        
        return {
            "success": True,
            "message": "æ˜ å°„åˆ é™¤æˆåŠŸ"
        }
    except Exception as e:
        db.rollback()
        logger.error(f"åˆ é™¤æ˜ å°„å¤±è´¥: {e}")
        return {"success": False, "message": str(e)}


@router.get("/export/mappings")
async def export_mappings(
    enabled: Optional[bool] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """
    å¯¼å‡ºåç§°æ˜ å°„ä¸ºExcelæ–‡ä»¶ï¼ˆåªåŒ…å«åŸåå’Œè‡ªå®šä¹‰åä¸¤åˆ—ï¼‰
    
    Args:
        enabled: è¿‡æ»¤å¯ç”¨çŠ¶æ€
        search: æœç´¢å…³é”®è¯
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        query = db.query(CustomNameMapping)
        
        if enabled is not None:
            query = query.filter(CustomNameMapping.enabled == enabled)
        
        if search:
            query = query.filter(
                (CustomNameMapping.original_name.like(f'%{search}%')) |
                (CustomNameMapping.quark_name.like(f'%{search}%')) |
                (CustomNameMapping.baidu_name.like(f'%{search}%')) |
                (CustomNameMapping.xunlei_name.like(f'%{search}%'))
            )
        
        mappings = query.order_by(CustomNameMapping.original_name).all()
        
        # åˆ›å»ºExcelå·¥ä½œç°¿
        wb = Workbook()
        ws = wb.active
        ws.title = "åç§°æ˜ å°„"
        
        # è®¾ç½®è¡¨å¤´
        headers = ["åŸå", "å¤¸å…‹æ˜¾ç¤ºå", "ç™¾åº¦æ˜¾ç¤ºå", "è¿…é›·æ˜¾ç¤ºå", "ç™¾åº¦é“¾æ¥", "å¤¸å…‹é“¾æ¥", "è¿…é›·é“¾æ¥"]
        ws.append(headers)
        
        # è®¾ç½®è¡¨å¤´æ ·å¼
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=14)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # å¡«å……æ•°æ®
        row_num = 2
        for mapping in mappings:
            ws.append([
                mapping.original_name,
                mapping.quark_name or '-',
                mapping.baidu_name or '-',
                mapping.xunlei_name or '-',
                mapping.baidu_link or '-',
                mapping.quark_link or '-',
                mapping.xunlei_link or '-'
            ])
            
            # è®¾ç½®å¯¹é½å’Œå­—ä½“
            for col_num in range(1, 8):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = Font(size=12)
            
            row_num += 1
        
        # è°ƒæ•´åˆ—å®½
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 60
        ws.column_dimensions['F'].width = 60
        ws.column_dimensions['G'].width = 60
        
        # æ·»åŠ ç»Ÿè®¡ä¿¡æ¯
        ws.append([])
        ws.append([f"å¯¼å‡ºæ—¶é—´: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws.append([f"æ€»è®¡: {len(mappings)} æ¡æ˜ å°„"])
        
        # ä¿å­˜åˆ°å†…å­˜
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # ç”Ÿæˆæ–‡ä»¶å
        filename = f"name_mappings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        logger.info(f"âœ… æˆåŠŸå¯¼å‡ºåç§°æ˜ å°„: {len(mappings)} æ¡")
        
        # è¿”å›Excelæ–‡ä»¶
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except ImportError:
        logger.error("openpyxlåº“æœªå®‰è£…ï¼Œæ— æ³•å¯¼å‡ºExcel")
        return {
            "success": False,
            "message": "openpyxlåº“æœªå®‰è£…ï¼Œè¯·å®‰è£…: pip install openpyxl"
        }
    except Exception as e:
        logger.error(f"å¯¼å‡ºå¤±è´¥: {e}")
        return {
            "success": False,
            "message": f"å¯¼å‡ºå¤±è´¥: {str(e)}"
        }


def process_anti_ban(target_dir: Path) -> int:
    """
    é˜²å¤±æ•ˆå¤„ç†ï¼šä¿®æ”¹ç›®å½•ä¸‹æ‰€æœ‰è§†é¢‘æ–‡ä»¶çš„Hash
    
    Args:
        target_dir: ç›®æ ‡ç›®å½•è·¯å¾„
        
    Returns:
        å¤„ç†çš„æ–‡ä»¶æ•°é‡
    """
    from backend.utils.obfuscator import FolderObfuscator
    
    processed_count = 0
    
    # éå†ç›®å½•ä¸‹æ‰€æœ‰æ–‡ä»¶
    for file_path in target_dir.rglob('*'):
        if not file_path.is_file():
            continue
        
        # åªå¤„ç†è§†é¢‘æ–‡ä»¶
        if not FolderObfuscator.is_video_file(str(file_path)):
            continue
        
        # æ£€æŸ¥æ˜¯å¦ä¸ºç¡¬é“¾æ¥
        stat_info = os.stat(file_path)
        if stat_info.st_nlink <= 1:
            # ä¸æ˜¯ç¡¬é“¾æ¥ï¼Œè·³è¿‡
            logger.info(f"è·³è¿‡éç¡¬é“¾æ¥æ–‡ä»¶: {file_path.name}")
            continue
        
        # ç”Ÿæˆä¸´æ—¶æ–‡ä»¶è·¯å¾„ï¼ˆä¿ç•™æ‰©å±•åï¼Œè®©ffmpegèƒ½è¯†åˆ«æ ¼å¼ï¼‰
        temp_file = file_path.parent / f"{file_path.stem}.tmp{file_path.suffix}"
        
        try:
            # ä½¿ç”¨ffmpegä¿®æ”¹Hash
            cmd = [
                'ffmpeg',
                '-i', str(file_path),
                '-map', '0',
                '-c', 'copy',
                '-metadata', f'comment={uuid.uuid4()}',
                '-metadata', f'title={uuid.uuid4()}',
                str(temp_file),
                '-y'  # è¦†ç›–å·²å­˜åœ¨çš„æ–‡ä»¶
            ]
            
            logger.info(f"æ­£åœ¨å¤„ç†: {file_path.name}")
            
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=300  # 5åˆ†é’Ÿè¶…æ—¶
            )
            
            if result.returncode != 0:
                logger.error(f"ffmpegå¤„ç†å¤±è´¥ {file_path.name}: {result.stderr}")
                # æ¸…ç†ä¸´æ—¶æ–‡ä»¶
                if temp_file.exists():
                    temp_file.unlink()
                continue
            
            # åˆ é™¤åŸç¡¬é“¾æ¥
            file_path.unlink()
            
            # é‡å‘½åä¸´æ—¶æ–‡ä»¶
            temp_file.rename(file_path)
            
            processed_count += 1
            logger.info(f"é˜²å¤±æ•ˆå¤„ç†å®Œæˆ: {file_path.name}")
            
        except subprocess.TimeoutExpired:
            logger.error(f"å¤„ç†è¶…æ—¶: {file_path.name}")
            if temp_file.exists():
                temp_file.unlink()
        except Exception as e:
            logger.error(f"å¤„ç†å¤±è´¥ {file_path.name}: {e}")
            if temp_file.exists():
                temp_file.unlink()
    
    return processed_count


class ResyncRequest(BaseModel):
    """é‡è½¬è¯·æ±‚"""
    original_name: str
    target_type: str  # 'quark', 'baidu' æˆ– 'xunlei'


def get_pinyin_initials(text: str) -> str:
    """
    æå–ä¸­æ–‡æ–‡æœ¬ç¬¬ä¸€ä¸ªå­—çš„æ‹¼éŸ³é¦–å­—æ¯
    
    Args:
        text: ä¸­æ–‡æ–‡æœ¬
        
    Returns:
        å¤§å†™æ‹¼éŸ³é¦–å­—æ¯ï¼ˆå•ä¸ªå­—æ¯ï¼‰
    """
    import re
    from backend.utils.pinyin_map import PINYIN_MAP
    
    # åªæå–ä¸­æ–‡å­—ç¬¦
    chinese_chars = re.findall(r'[\u4e00-\u9fff]', text)
    if not chinese_chars:
        return ""
    
    # åªå–ç¬¬ä¸€ä¸ªå­—
    first_char = chinese_chars[0]
    pinyin = PINYIN_MAP.get(first_char, '')
    if pinyin:
        return pinyin[0].upper()
    
    return ""


@router.post("/mappings/obfuscate")
async def obfuscate_name(
    original_name: str,
    db: Session = Depends(get_db)
):
    """
    æ··æ·†åç§°æ¥å£ - è¿”å›æ··æ·†åçš„åç§°
    
    å·²å­˜åœ¨çš„æ˜ å°„ï¼šè¿”å›æ•°æ®åº“ä¸­çš„åç§°ï¼ˆä¿æŒåŸæ ·ï¼‰
    æ–°å»ºçš„æ˜ å°„ï¼šåœ¨æ··æ·†åç§°å‰åŠ é¦–å­—æ¯ç¼©å†™
    
    Args:
        original_name: åŸå§‹åç§°
        
    Returns:
        æ··æ·†åçš„åç§°
    """
    try:
        from backend.main import db_engine
        import re
        
        # 1. æ£€æŸ¥æ•°æ®åº“æ˜¯å¦å·²å­˜åœ¨è¯¥æ˜ å°„
        existing = db.query(CustomNameMapping).filter(
            CustomNameMapping.original_name == original_name
        ).first()
        
        if existing:
            # å·²å­˜åœ¨ï¼šè¿”å›æ•°æ®åº“ä¸­çš„ä¸‰ä¸ªåç§°
            logger.info(f"âœ… ä½¿ç”¨å·²å­˜åœ¨æ˜ å°„: {original_name}")
            return {
                "success": True,
                "data": {
                    "original_name": original_name,
                    "quark_name": existing.quark_name or "",
                    "baidu_name": existing.baidu_name or "",
                    "xunlei_name": existing.xunlei_name or "",
                    "is_existing": True
                }
            }
        
        # 2. æ–°å»ºçš„æ˜ å°„ï¼šè°ƒç”¨å…¬å…±æ··æ·†æ–¹æ³•ï¼ˆä¸ä¿å­˜åˆ°æ•°æ®åº“ï¼Œåªè¿”å›æ··æ·†ç»“æœï¼‰
        from backend.utils.obfuscator import FolderObfuscator
        obfuscator = FolderObfuscator(enabled=True, db_engine=db_engine)
        
        # ä½¿ç”¨å…¬å…±æ–¹æ³•ï¼šå»å¹´ä»½ + åŒéŸ³å­—æ··æ·† + æ·»åŠ é¦–å­—æ¯
        obfuscated_name = obfuscator.obfuscate_with_initial(original_name)
        
        # æå–é¦–å­—æ¯ç”¨äºè¿”å›ï¼ˆå‰ç«¯å¯èƒ½éœ€è¦ï¼‰
        base_name = re.sub(r'\s*\((\d{4})\)\s*$', '', original_name).strip()
        initials = obfuscated_name.split()[0] if ' ' in obfuscated_name else ""
        
        logger.info(f"âœ… æ··æ·†åç§°(æ–°å»º): {original_name} -> {obfuscated_name}")
        
        return {
            "success": True,
            "data": {
                "original_name": original_name,
                "obfuscated_name": obfuscated_name,
                "is_existing": False,
                "initials": initials
            }
        }
    except Exception as e:
        logger.error(f"æ··æ·†åç§°å¤±è´¥: {e}")
        import traceback
        traceback.print_exc()
        return {
            "success": False,
            "message": f"æ··æ·†å¤±è´¥: {str(e)}"
        }


def _do_resync_task(
    original_name: str,
    target_type: str
):
    """
    åå°ä»»åŠ¡ï¼šæ‰§è¡Œé‡è½¬æ“ä½œ
    """
    from backend.main import db_engine
    from backend.models import get_session
    
    db = get_session(db_engine)
    try:
        from pathlib import Path
        import shutil
        from backend.utils.linker import FileLinker
        from backend.main import db_engine, config
        
        # æŸ¥è¯¢æ˜ å°„è·å–æ˜¾ç¤ºåç§°
        mapping = db.query(CustomNameMapping).filter(
            CustomNameMapping.original_name == original_name
        ).first()
        
        if not mapping:
            logger.error(f"æœªæ‰¾åˆ°æ˜ å°„: {original_name}")
            return
        
        # æ¸…ç©ºå¯¹åº”ç½‘ç›˜çš„é“¾æ¥å­—æ®µ
        if target_type == 'quark':
            mapping.quark_link = None
            logger.info(f"âœ… å·²æ¸…ç©º '{original_name}' çš„å¤¸å…‹é“¾æ¥")
        elif target_type == 'baidu':
            mapping.baidu_link = None
            logger.info(f"âœ… å·²æ¸…ç©º '{original_name}' çš„ç™¾åº¦é“¾æ¥")
        elif target_type == 'xunlei':
            mapping.xunlei_link = None
            logger.info(f"âœ… å·²æ¸…ç©º '{original_name}' çš„è¿…é›·é“¾æ¥")
        db.commit()
        
        # è·å–æ˜¾ç¤ºåç§°
        if target_type == 'quark':
            display_name = mapping.quark_name
        elif target_type == 'baidu':
            display_name = mapping.baidu_name
        elif target_type == 'xunlei':
            display_name = mapping.xunlei_name
        else:
            display_name = None
        
        if not display_name:
            display_name = original_name
        
        # æŸ¥è¯¢æ‰€æœ‰è¯¥å‰§é›†çš„è®°å½•
        records = db.query(LinkRecord).filter(
            LinkRecord.original_name == original_name
        ).all()
        
        if not records:
            logger.error(f"æœªæ‰¾åˆ°åŒæ­¥è®°å½•: {original_name}")
            return
        
        # è·å–é…ç½®
        monitors = config.get('monitors', [])
        if not monitors:
            return {"success": False, "message": "æœªæ‰¾åˆ°ç›‘æ§é…ç½®"}
        
        monitor = monitors[0]
        source_path = Path(monitor.get('source', ''))
        targets = monitor.get('targets', [])
        
        if len(targets) < 2:
            return {"success": False, "message": "ç›®æ ‡é…ç½®ä¸è¶³"}
        
        # ç¡®å®šç›®æ ‡è·¯å¾„
        if target_type == 'quark':
            target_idx = 0
        elif target_type == 'baidu':
            target_idx = 1
        elif target_type == 'xunlei':
            target_idx = 2
        else:
            logger.error(f"ä¸æ”¯æŒçš„ç½‘ç›˜ç±»å‹: {target_type}")
            return
        
        if len(targets) <= target_idx:
            return {"success": False, "message": f"ç›®æ ‡é…ç½®ä¸è¶³ï¼Œéœ€è¦è‡³å°‘{target_idx + 1}ä¸ªç›®æ ‡"}
        
        target_config = targets[target_idx]
        target_base = Path(target_config.get('path', ''))
        
        # åˆå§‹åŒ–linker
        obfuscate_enabled = monitor.get('obfuscate_enabled', True)
        linker = FileLinker(obfuscate_enabled=obfuscate_enabled, db_engine=db_engine)
        
        # ä¸´æ—¶æ›´æ–°æ˜ å°„è¡¨ï¼Œè®©æ··æ·†å™¨ä½¿ç”¨æ–°åç§°
        if target_type == 'quark':
            old_name = mapping.quark_name
            mapping.quark_name = display_name
        elif target_type == 'baidu':
            old_name = mapping.baidu_name
            mapping.baidu_name = display_name
        elif target_type == 'xunlei':
            old_name = mapping.xunlei_name
            mapping.xunlei_name = display_name
        db.commit()
        
        success_count = 0
        failed_count = 0
        deleted_dirs = set()
        
        for record in records:
            try:
                source_file = Path(record.source_file)
                if not source_file.exists():
                    logger.warning(f"æºæ–‡ä»¶ä¸å­˜åœ¨: {source_file}")
                    continue
                
                # è·å–æ—§çš„ç›®æ ‡æ–‡ä»¶è·¯å¾„
                if target_type == 'quark':
                    old_target_file = record.quark_target_file
                elif target_type == 'baidu':
                    old_target_file = record.baidu_target_file
                elif target_type == 'xunlei':
                    old_target_file = record.xunlei_target_file
                else:
                    old_target_file = None
                
                # åˆ é™¤æ—§æ–‡ä»¶
                if old_target_file:
                    old_path = Path(old_target_file)
                    if old_path.exists():
                        old_path.unlink()
                        logger.info(f"å·²åˆ é™¤æ—§æ–‡ä»¶: {old_path}")
                
                # è®¡ç®—æ–°çš„ç›®æ ‡è·¯å¾„
                relative_path = source_file.relative_to(source_path)
                new_target_file = target_base / relative_path
                
                # é‡æ–°ç¡¬é“¾æ¥
                success, method, error, actual_target = linker.create_hardlink(
                    source_file, new_target_file,
                    source_base=source_path,
                    target_base=target_base
                )
                
                if success and actual_target:
                    # æ›´æ–°è®°å½•
                    if target_type == 'quark':
                        record.quark_target_file = str(actual_target)
                        record.quark_synced_at = datetime.now()
                    elif target_type == 'baidu':
                        record.baidu_target_file = str(actual_target)
                        record.baidu_synced_at = datetime.now()
                    elif target_type == 'xunlei':
                        record.xunlei_target_file = str(actual_target)
                        record.xunlei_synced_at = datetime.now()
                    
                    record.updated_at = datetime.now()
                    success_count += 1
                    logger.info(f"é‡è½¬æˆåŠŸ: {source_file.name} -> {actual_target}")
                    
                    # å¦‚æœæ—§æ–‡ä»¶å­˜åœ¨ä¸”æ–°æ—§ç›®å½•ä¸åŒï¼Œæ”¶é›†æ—§ç›®å½•ç”¨äºåˆ é™¤
                    if old_target_file:
                        old_show_dir = Path(old_target_file).parent.parent
                        new_show_dir = Path(actual_target).parent.parent
                        if old_show_dir != new_show_dir and old_show_dir.exists():
                            deleted_dirs.add(old_show_dir)
                else:
                    failed_count += 1
                    logger.error(f"é‡è½¬å¤±è´¥: {source_file.name}, é”™è¯¯: {error}")
                    
            except Exception as e:
                failed_count += 1
                logger.error(f"å¤„ç†æ–‡ä»¶å¤±è´¥ {record.source_file}: {e}")
        
        # åˆ é™¤ç©ºç›®å½•ï¼ˆåªåˆ é™¤çœŸæ­£ä¸ºç©ºçš„ç›®å½•ï¼‰
        for dir_path in deleted_dirs:
            if dir_path.exists():
                try:
                    # æ£€æŸ¥ç›®å½•æ˜¯å¦ä¸ºç©º
                    if not any(dir_path.iterdir()):
                        shutil.rmtree(dir_path)
                        logger.info(f"å·²åˆ é™¤ç©ºç›®å½•: {dir_path}")
                    else:
                        logger.info(f"è·³è¿‡éç©ºç›®å½•: {dir_path}")
                except Exception as e:
                    logger.error(f"åˆ é™¤ç›®å½•å¤±è´¥ {dir_path}: {e}")
        
        db.commit()
        
        # é˜²å¤±æ•ˆå¤„ç†ï¼šä¿®æ”¹è§†é¢‘æ–‡ä»¶Hash
        if success_count > 0:
            try:
                # ä»å·²é‡è½¬çš„æ–‡ä»¶è®°å½•ä¸­æå–å®é™…çš„å‰§é›†æ ¹ç›®å½•
                sample_record = db.query(LinkRecord).filter(
                    LinkRecord.original_name == original_name
                ).first()
                
                if sample_record:
                    # æ ¹æ®target_typeè·å–å¯¹åº”çš„target_file
                    target_file_str = None
                    if target_type == 'quark':
                        target_file_str = sample_record.quark_target_file
                    elif target_type == 'baidu':
                        target_file_str = sample_record.baidu_target_file
                    elif target_type == 'xunlei':
                        target_file_str = sample_record.xunlei_target_file
                    
                    if target_file_str:
                        # ä¾‹å¦‚: /media/å¤¸å…‹ç½‘ç›˜/å‰§é›†/å›½äº§å‰§é›†/Cé’é’æ„Ÿè´¡ç±³(2025)/Season 1/S01E01.mkv
                        # æå–å‰§é›†æ ¹ç›®å½•: /media/å¤¸å…‹ç½‘ç›˜/å‰§é›†/å›½äº§å‰§é›†/Cé’é’æ„Ÿè´¡ç±³(2025)
                        target_file_path = Path(target_file_str)
                        
                        # å‘ä¸Šæ‰¾2çº§ï¼šæ–‡ä»¶ -> Season 1 -> å‰§é›†æ ¹ç›®å½•
                        if target_file_path.exists():
                            target_show_dir = target_file_path.parent.parent
                            
                            logger.info(f"å¼€å§‹é˜²å¤±æ•ˆå¤„ç†: {target_show_dir}")
                            processed_count = process_anti_ban(target_show_dir)
                            logger.info(f"é˜²å¤±æ•ˆå¤„ç†å®Œæˆ: å¤„ç†äº† {processed_count} ä¸ªè§†é¢‘æ–‡ä»¶")
                        else:
                            logger.warning(f"ç›®æ ‡æ–‡ä»¶ä¸å­˜åœ¨: {target_file_path}")
                    else:
                        logger.warning(f"æœªæ‰¾åˆ°ç›®æ ‡æ–‡ä»¶è·¯å¾„")
                else:
                    logger.warning(f"æœªæ‰¾åˆ°é‡è½¬è®°å½•")
            except Exception as e:
                logger.error(f"é˜²å¤±æ•ˆå¤„ç†å¤±è´¥: {e}")
                import traceback
                traceback.print_exc()
        
        # è§¦å‘TaoSyncåŒæ­¥
        if success_count > 0:
            try:
                from backend.main import monitor_service
                if monitor_service and monitor_service.handlers:
                    for handler in monitor_service.handlers:
                        if hasattr(handler, 'taosync_queue') and handler.taosync_queue:
                            logger.info(f"é‡è½¬å®Œæˆï¼Œè§¦å‘TaoSyncåŒæ­¥ï¼ˆ{success_count}ä¸ªæ–‡ä»¶ï¼‰")
                            handler.taosync_queue.trigger_now(file_count=success_count)
                            break
            except Exception as e:
                logger.error(f"è§¦å‘TaoSyncå¤±è´¥: {e}")
        
        target_names = {'quark': 'å¤¸å…‹', 'baidu': 'ç™¾åº¦', 'xunlei': 'è¿…é›·'}
        target_name = target_names.get(target_type, 'æœªçŸ¥')
        logger.info(f"âœ… é‡è½¬å®Œæˆ: {original_name} -> {target_name}, æˆåŠŸ {success_count} ä¸ªï¼Œå¤±è´¥ {failed_count} ä¸ª")
        
    except Exception as e:
        logger.error(f"é‡è½¬ä»»åŠ¡å¼‚å¸¸: {e}")
        db.rollback()
        import traceback
        traceback.print_exc()
    finally:
        db.close()


@router.post("/mappings/resync")
async def resync_to_target(
    request: ResyncRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db)
):
    """
    é‡è½¬å‰§é›†åˆ°æŒ‡å®šç½‘ç›˜ï¼ˆåå°ä»»åŠ¡ï¼‰
    
    1. ç«‹å³è¿”å›å“åº”ï¼Œä¸é˜»å¡è¯·æ±‚
    2. åœ¨åå°æ‰§è¡Œé‡è½¬æ“ä½œ
    3. å‰ç«¯å¯ä»¥ç»§ç»­æ“ä½œï¼Œä¸ä¼šå¡æ­»
    
    æ“ä½œæµç¨‹ï¼š
    - æŸ¥è¯¢è¯¥å‰§é›†çš„æ‰€æœ‰æ–‡ä»¶è®°å½•
    - æŸ¥è¯¢æ˜ å°„è¡¨è·å–æ˜¾ç¤ºåç§°
    - åˆ é™¤ç›®æ ‡ç½‘ç›˜çš„æ—§æ–‡ä»¶
    - ç”¨æ–°åç§°é‡æ–°åˆ›å»ºç¡¬é“¾æ¥
    - è§¦å‘é˜²å¤±æ•ˆå¤„ç†å’ŒTaoSyncåŒæ­¥
    """
    # éªŒè¯æ˜ å°„æ˜¯å¦å­˜åœ¨
    mapping = db.query(CustomNameMapping).filter(
        CustomNameMapping.original_name == request.original_name
    ).first()
    
    if not mapping:
        raise HTTPException(
            status_code=404,
            detail=f"æœªæ‰¾åˆ° '{request.original_name}' çš„æ˜ å°„é…ç½®"
        )
    
    # éªŒè¯æ˜¯å¦æœ‰è®°å½•
    records_count = db.query(LinkRecord).filter(
        LinkRecord.original_name == request.original_name
    ).count()
    
    if records_count == 0:
        raise HTTPException(
            status_code=404,
            detail=f"æœªæ‰¾åˆ° '{request.original_name}' çš„åŒæ­¥è®°å½•"
        )
    
    # æ·»åŠ åå°ä»»åŠ¡
    background_tasks.add_task(
        _do_resync_task,
        request.original_name,
        request.target_type
    )
    
    target_names = {'quark': 'å¤¸å…‹', 'baidu': 'ç™¾åº¦', 'xunlei': 'è¿…é›·'}
    target_name = target_names.get(request.target_type, 'æœªçŸ¥')
    
    logger.info(f"ğŸš€ é‡è½¬ä»»åŠ¡å·²å¯åŠ¨: {request.original_name} -> {target_name} ({records_count}ä¸ªæ–‡ä»¶)")
    
    return {
        "success": True,
        "message": f"é‡è½¬ä»»åŠ¡å·²å¯åŠ¨ï¼Œæ­£åœ¨åå°å¤„ç† {records_count} ä¸ªæ–‡ä»¶...",
        "data": {
            "target_type": target_name,
            "file_count": records_count,
            "original_name": request.original_name
        }
    }
